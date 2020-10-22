from tkinter import Tk, Label, Button, Entry, Frame, PhotoImage, StringVar
import datetime
from os import path
from openpyxl import load_workbook


class Worker:
    
    hours = None
    start_minutes = None
    end_minutes = None
    start_timer = None
    end_timer = None
    start = True
    current_date = None
    current_time_start = None
    current_time_end = None

    def __init__(self, name):
        self.name = name
        if path.exists(f'{self.name} timer.xlsx'):
            self.start_work()
        else:
            raise FileExistsError('Ligger ikke i systemet!')

    def start_work(self):
        self.start = False
        self.hours = 0
        date = datetime.datetime.now()
        self.current_time_start = date.strftime("%H:%M:%S")
        self.current_date = date.year, date.month, date.day
        self.start_timer = date.hour
        self.start_minutes = date.minute
        year = self.current_date[0]
        month = self.current_date[1]
        wb = load_workbook(f'{self.name} timer.xlsx')
        if f'Arbeid {month, year}' not in wb:
            wb.create_sheet(f'Arbeid {month, year}')
            ws = wb[f'Arbeid {month, year}']
            ws['A1'] = 'Start Tid'
            ws['B1'] = 'Slutt Tid'
            ws['C1'] = 'Dato'
            ws['D1'] = 'Timer'
            wb.save(f'{self.name} timer.xlsx')

        else:
            wb.save(f'{self.name} timer.xlsx')

    def end_work(self):
        if self.start == False:
            date = datetime.datetime.now()
            self.current_time_end = date.strftime("%H:%M:%S")
            self.end_timer = date.hour
            self.end_minutes = date.minute
            self.work_calculator()
            year = self.current_date[0]
            month = self.current_date[1]
            day = self.current_date[2]
            file_name = f'{self.name} timer.xlsx'
            sheet = f'Arbeid {month, year}'

            df = load_workbook(file_name)
            ws = df[sheet]
            df2 = [self.current_time_start, self.current_time_end, f'd{day}:m{month}:y{year}', self.hours]
            ws.append(df2)
            df.save(file_name)
            self.start = True
        else:
            return False

    def work_calculator(self):
        if self.end_timer < self.start_timer:
            timer = (24 + self.end_timer - self.start_timer)*60
            minutter = (self.end_minutes - self.start_minutes)
            self.hours = (timer + minutter)/60
        else:
            timer = (self.end_timer - self.start_timer)*60
            minutter = (self.end_minutes - self.start_minutes)
            self.hours = (timer + minutter)/60



root = Tk()

fyrer = {}
person_navn = StringVar()
person_navn.set('')


def start_work():
    try:
        person = enter_start.get()
        fyrer[person] = Worker(person)
        person_navn.set(f"{person} started work")
    except ValueError:
        return False


def end_work():
    try:
        person = enter_start.get()
        fyrer[person].end_work()
        person_navn.set(f"{person} ended work")
    except ValueError:
        return False


frame = Frame(root, bg='#80c1ff', bd=5)
frame.place(relx=0.5, rely=0.1, relwidth=1, relheight=1, anchor='n')

label_text = Label(frame, text='Enter name:', bg='#80c1ff')
label_text.pack()

enter_start = Entry(frame, font=40, justify='center')
enter_start.pack()

start_button = Button(frame, text='Start work', command=start_work)
start_button.pack()

end_button = Button(frame, text='End work', command=end_work)
end_button.pack()

myLabel = Label(frame, textvariable=person_navn, bg='#80c1ff')
myLabel.pack()


root.mainloop()